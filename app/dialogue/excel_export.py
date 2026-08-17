import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def calculate_dialogue_statistics(
        engine,
        dialogue_id
):

    move_counts = {
        "PROPOSE": 0,
        "SUPPORT": 0,
        "CHALLENGE": 0,
        "ACCEPT": 0,
        "REJECT": 0,
        "WITHDRAW": 0
    }

    burden_counts = {
        "CREATED": 0,
        "ACTIVATED": 0,
        "SATISFIED": 0,
        "RESOLVED": 0
    }

    for turn in engine.transcript.turns:

        move = turn.get("move")

        if move in move_counts:
            move_counts[move] += 1

        burden_event = turn.get(
            "burden_event"
        )

        if burden_event in burden_counts:
            burden_counts[burden_event] += 1

    termination_reason = (
        engine.termination_reason
        or "UNKNOWN"
    )

    accepted_flag = (
        1
        if termination_reason == "PROPOSAL_ACCEPTED"
        else 0
    )

    rejected_flag = (
        1
        if termination_reason == "PROPOSAL_REJECTED"
        else 0
    )

    accepted_proposal = None

    if accepted_flag == 1:
        accepted_proposal = engine.current_proposal

    return {
        "Dialogue ID": dialogue_id,
        "Protocol": engine.protocol,
        "Turn Count": len(
            engine.transcript.turns
        ),
        "Outcome": termination_reason,
        "Termination Reason": termination_reason,
        "Accepted Proposal": accepted_proposal,
        "PROPOSE": move_counts["PROPOSE"],
        "SUPPORT": move_counts["SUPPORT"],
        "CHALLENGE": move_counts["CHALLENGE"],
        "ACCEPT": move_counts["ACCEPT"],
        "REJECT": move_counts["REJECT"],
        "WITHDRAW": move_counts["WITHDRAW"],
        "Burdens Created": burden_counts["CREATED"],
        "Burdens Activated": burden_counts["ACTIVATED"],
        "Burdens Satisfied": burden_counts["SATISFIED"],
        "Burdens Resolved": burden_counts["RESOLVED"],
        "Accepted Flag": accepted_flag,
        "Rejected Flag": rejected_flag
    }


def get_next_results_filename():

    folder = "app/results"

    os.makedirs(
        folder,
        exist_ok=True
    )

    existing_files = [
        filename
        for filename in os.listdir(folder)
        if (
            filename.startswith(
                "simulation_results_"
            )
            and filename.endswith(
                ".xlsx"
            )
        )
    ]

    numbers = []

    for filename in existing_files:

        try:

            number = int(
                filename
                .replace(
                    "simulation_results_",
                    ""
                )
                .replace(
                    ".xlsx",
                    ""
                )
            )

            numbers.append(number)

        except ValueError:
            continue

    next_number = (
        max(numbers) + 1
        if numbers
        else 1
    )

    return os.path.join(
        folder,
        f"simulation_results_{next_number:03}.xlsx"
    )


def style_header(worksheet):

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


def autofit_columns(worksheet):

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            value = cell.value

            if value is None:
                continue

            value_length = len(
                str(value)
            )

            if value_length > max_length:
                max_length = value_length

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            28
        )


def create_raw_data_sheet(
        workbook,
        rows
):

    worksheet = workbook.active

    worksheet.title = "Raw Data"

    headers = [
        "Dialogue ID",
        "Protocol",
        "Turn Count",
        "Outcome",
        "Termination Reason",
        "Accepted Proposal",
        "PROPOSE",
        "SUPPORT",
        "CHALLENGE",
        "ACCEPT",
        "REJECT",
        "WITHDRAW",
        "Burdens Created",
        "Burdens Activated",
        "Burdens Satisfied",
        "Burdens Resolved",
        "Accepted Flag",
        "Rejected Flag"
    ]

    worksheet.append(headers)

    for row in rows:

        worksheet.append([
            row[header]
            for header in headers
        ])

    style_header(worksheet)

    worksheet.freeze_panes = "A2"

    autofit_columns(worksheet)

    return worksheet


def get_protocol_rows(
        rows,
        protocol
):

    return [
        row
        for row in rows
        if row["Protocol"] == protocol
    ]


def average(values):

    if not values:
        return 0

    return sum(values) / len(values)


def create_protocol_summary_sheet(
        workbook,
        rows
):

    worksheet = workbook.create_sheet(
        "Protocol Summary"
    )

    headers = [
        "Protocol",
        "Dialogues",
        "Average Turns",
        "Minimum Turns",
        "Maximum Turns",
        "Acceptance Rate",
        "Rejection Rate",
        "Average PROPOSE",
        "Average SUPPORT",
        "Average CHALLENGE",
        "Average ACCEPT",
        "Average REJECT",
        "Average WITHDRAW",
        "Burdens Created",
        "Burdens Activated",
        "Burdens Satisfied",
        "Burdens Resolved",
        "Burden Satisfaction Rate"
    ]

    worksheet.append(headers)

    for protocol in [
        "Standard",
        "Burden"
    ]:

        protocol_rows = get_protocol_rows(
            rows,
            protocol
        )

        dialogue_count = len(
            protocol_rows
        )

        turn_counts = [
            row["Turn Count"]
            for row in protocol_rows
        ]

        accepted_count = sum(
            row["Accepted Flag"]
            for row in protocol_rows
        )

        rejected_count = sum(
            row["Rejected Flag"]
            for row in protocol_rows
        )

        burdens_created = sum(
            row["Burdens Created"]
            for row in protocol_rows
        )

        burdens_activated = sum(
            row["Burdens Activated"]
            for row in protocol_rows
        )

        burdens_satisfied = sum(
            row["Burdens Satisfied"]
            for row in protocol_rows
        )

        burdens_resolved = sum(
            row["Burdens Resolved"]
            for row in protocol_rows
        )

        acceptance_rate = (
            accepted_count / dialogue_count
            if dialogue_count
            else 0
        )

        rejection_rate = (
            rejected_count / dialogue_count
            if dialogue_count
            else 0
        )

        burden_satisfaction_rate = (
            burdens_satisfied / burdens_activated
            if burdens_activated
            else 0
        )

        worksheet.append([
            protocol,
            dialogue_count,
            round(
                average(turn_counts),
                2
            ),
            min(turn_counts)
            if turn_counts
            else 0,
            max(turn_counts)
            if turn_counts
            else 0,
            acceptance_rate,
            rejection_rate,
            round(
                average([
                    row["PROPOSE"]
                    for row in protocol_rows
                ]),
                2
            ),
            round(
                average([
                    row["SUPPORT"]
                    for row in protocol_rows
                ]),
                2
            ),
            round(
                average([
                    row["CHALLENGE"]
                    for row in protocol_rows
                ]),
                2
            ),
            round(
                average([
                    row["ACCEPT"]
                    for row in protocol_rows
                ]),
                2
            ),
            round(
                average([
                    row["REJECT"]
                    for row in protocol_rows
                ]),
                2
            ),
            round(
                average([
                    row["WITHDRAW"]
                    for row in protocol_rows
                ]),
                2
            ),
            burdens_created,
            burdens_activated,
            burdens_satisfied,
            burdens_resolved,
            burden_satisfaction_rate
        ])

    style_header(worksheet)

    worksheet.freeze_panes = "A2"

    # Percentage formatting
    for cell in [
        "F2", "F3",
        "G2", "G3",
        "R2", "R3"
    ]:

        worksheet[cell].number_format = (
            "0.00%"
        )

    autofit_columns(worksheet)

    return worksheet


def create_termination_summary_sheet(
        workbook,
        rows
):

    worksheet = workbook.create_sheet(
        "Termination Summary"
    )

    worksheet.append([
        "Protocol",
        "Termination Reason",
        "Count",
        "Percentage"
    ])

    protocols = [
        "Standard",
        "Burden"
    ]

    termination_reasons = sorted(
        {
            row["Termination Reason"]
            for row in rows
        }
    )

    for protocol in protocols:

        protocol_rows = get_protocol_rows(
            rows,
            protocol
        )

        total = len(
            protocol_rows
        )

        for reason in termination_reasons:

            count = sum(
                1
                for row in protocol_rows
                if (
                    row[
                        "Termination Reason"
                    ]
                    == reason
                )
            )

            percentage = (
                count / total
                if total
                else 0
            )

            worksheet.append([
                protocol,
                reason,
                count,
                percentage
            ])

    style_header(worksheet)

    for row_number in range(
        2,
        worksheet.max_row + 1
    ):

        worksheet[
            f"D{row_number}"
        ].number_format = "0.00%"

    autofit_columns(worksheet)

    return worksheet


def export_comparative_results(
        standard_engines,
        burden_engines
):

    workbook = Workbook()

    rows = []

    dialogue_id = 1

    for engine in standard_engines:

        rows.append(
            calculate_dialogue_statistics(
                engine,
                dialogue_id
            )
        )

        dialogue_id += 1

    for engine in burden_engines:

        rows.append(
            calculate_dialogue_statistics(
                engine,
                dialogue_id
            )
        )

        dialogue_id += 1

    create_raw_data_sheet(
        workbook,
        rows
    )

    create_protocol_summary_sheet(
        workbook,
        rows
    )

    create_termination_summary_sheet(
        workbook,
        rows
    )

    output_path = (
        get_next_results_filename()
    )

    workbook.save(
        output_path
    )

    print()
    print(
        "Excel results exported successfully!"
    )

    print(
        f"File: {output_path}"
    )

    return output_path